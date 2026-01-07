import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import io

# --- Configuration de la Page ---
st.set_page_config(page_title="Mise à jour Prix (Comparaison)", page_icon="⚖️")

st.title("⚖️ Mise à Jour : Ancien vs Nouveau")
st.markdown("""
**Fonctionnement :**
Cette version ne remplace pas vos prix ! 
Elle écrit le **Nouveau Prix** dans la colonne **juste à droite** de l'ancien prix.

1. **Fichier Cible** : Votre fichier actuel (ex: `PCI 2026`).
2. **Fichier Source** : Le fichier avec les nouveaux prix (ex: `PCI 2024`).
3. Résultat : Vous aurez `Prix Actuel` | `Prix Source` côte à côte.
""")

# --- Fonctions Utilitaires ---

def clean_code(val):
    """Nettoie le code produit pour la comparaison"""
    if val is None: return ""
    return str(val).strip().upper()

def get_source_prices(file_source):
    """Lit le fichier source avec Pandas pour extraire {Code: Prix}"""
    # 1. Essayer avec header=0 puis header=1
    try:
        df = pd.read_excel(file_source, header=0)
        cols = [str(c).upper() for c in df.columns]
        if not any("CODE" in c for c in cols):
            df = pd.read_excel(file_source, header=1)
    except:
        return None, "Erreur lecture fichier source"

    # 2. Trouver les colonnes
    col_code = None
    col_price = None

    for col in df.columns:
        c_str = str(col).upper()
        if "CODE" in c_str or "NOMENCLATURE" in c_str:
            col_code = col
        if "PCI" in c_str or "PRIX" in c_str or "PRICE" in c_str:
            # Priorité à "PCI CAISSE" ou "PCI PCI"
            if col_price is None or "CAISSE" in c_str:
                col_price = col
    
    if not col_code or not col_price:
        return None, f"Colonnes introuvables (Source). Colonnes détectées : {list(df.columns)}"

    # 3. Créer le dictionnaire
    price_dict = {}
    for _, row in df.iterrows():
        code = clean_code(row[col_code])
        price = row[col_price]
        if pd.notna(price) and isinstance(price, (int, float)):
            price_dict[code] = round(price, 2)
            
    return price_dict, None

def update_excel_side_by_side(file_target, price_dict):
    """Ouvre le fichier cible et écrit le nouveau prix À DROITE de l'ancien"""
    wb = load_workbook(file_target)
    ws = wb.active

    # 1. Scanner l'en-tête (Lignes 1-5)
    header_row_idx = None
    col_map = {}

    for r in range(1, 6):
        row_values = [str(cell.value).upper() if cell.value else "" for cell in ws[r]]
        if any("CODE" in s for s in row_values) and (any("PCI" in s for s in row_values) or any("PRIX" in s for s in row_values)):
            header_row_idx = r
            for i, val in enumerate(row_values):
                # On ne mappe que si la cellule n'est pas vide
                if val: col_map[val] = i + 1 
            break
    
    if header_row_idx is None:
        return None, "Impossible de trouver l'en-tête (Code/Prix) dans le fichier Cible."

    # 2. Identifier les colonnes
    idx_code = None
    idx_price_target = None

    for name, idx in col_map.items():
        if "CODE" in name or "NOMENCLATURE" in name:
            idx_code = idx
        if "PCI" in name or "PRIX" in name:
             if idx_price_target is None or "CAISSE" in name or "PCI PCI" in name:
                idx_price_target = idx

    if not idx_code or not idx_price_target:
        return None, "Colonnes cibles non identifiées."

    # Colonne de destination = Colonne Prix + 1 (Juste à droite)
    idx_dest = idx_price_target + 1

    # 3. Ajouter un En-tête pour la nouvelle colonne
    header_cell = ws.cell(row=header_row_idx, column=idx_dest)
    header_cell.value = "Prix Source (Nouveau)"
    header_cell.font = Font(bold=True, color="FF0000") # En rouge pour être visible

    # 4. Remplir les prix
    count = 0
    # Parcourir les lignes
    for r in range(header_row_idx + 1, ws.max_row + 1):
        cell_code = ws.cell(row=r, column=idx_code)
        
        # On lit le code
        code_val = clean_code(cell_code.value)
        
        if code_val in price_dict:
            new_price = price_dict[code_val]
            
            # On écrit dans la colonne de destination (à droite)
            cell_dest = ws.cell(row=r, column=idx_dest)
            cell_dest.value = new_price
            count += 1

    # Sauvegarde
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer, count

# --- Interface Utilisateur ---
col1, col2 = st.columns(2)

with col1:
    f_target = st.file_uploader("📝 Fichier Cible (À compléter)", type=['xlsx'])

with col2:
    f_source = st.file_uploader("💰 Fichier Source (Nouveaux Prix)", type=['xlsx'])

if f_target and f_source:
    if st.button("Comparer Côte à Côte 🚀"):
        with st.spinner("Traitement..."):
            
            prices, error = get_source_prices(f_source)
            
            if error:
                st.error(f"Erreur Source : {error}")
            else:
                st.info(f"{len(prices)} prix trouvés dans la source.")

                result_buffer, count_or_err = update_excel_side_by_side(f_target, prices)

                if isinstance(count_or_err, str):
                    st.error(f"Erreur Cible : {count_or_err}")
                else:
                    st.success(f"✅ Terminé ! {count_or_err} prix ajoutés dans la colonne à droite.")
                    
                    st.download_button(
                        label="📥 Télécharger le Comparatif",
                        data=result_buffer,
                        file_name="Comparatif_Prix_Cote_a_Cote.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ): {e}")
